import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import requests
import streamlit as st


REQUEST_TIMEOUT = 60
DEFAULT_LANGUAGE = "en-US"


# --------- REGION-BASED ENDPOINTS ---------
def get_endpoints(region: str) -> Tuple[str, str]:
    """Return auth endpoint and API base (with /api/v1) for the selected region."""
    if region == "North America":
        auth = "https://us-auth.hammertechonline.com/api/login/generatetoken"
        api_base = "https://us-api.hammertechonline.com/api/v1"
    elif region == "Asia/Australia/NZ":
        auth = "https://au-auth.hammertechonline.com/api/login/generatetoken"
        api_base = "https://au-api.hammertechonline.com/api/v1"
    else:  # Europe/UK
        auth = "https://eu-auth.hammertechonline.com/api/login/generatetoken"
        api_base = "https://eu-api.hammertechonline.com/api/v1"

    return auth, api_base


SHEET_CONFIG: Dict[str, Dict[str, str]] = {
    "Users": {
        "label": "Users",
        "endpoint_key": "users",
        "description": (
            "Expected columns: Email Address, Full Name, Phone, Job Title, "
            "Internal Identifier, Demo Project ID"
        ),
    },
    "Projects": {
        "label": "Projects",
        "endpoint_key": "projects",
        "description": (
            "Expected columns: Project Name, Country, Site Address, Time Zone String, "
            "State, Internal ID, Region ID"
        ),
    },
    "EmployerProfiles": {
        "label": "Employer Profiles",
        "endpoint_key": "employer_profiles",
        "description": (
            "Expected columns: Business Name, ABN, Street Address, City / Suburb, "
            "State / Province, Postal Code, Country, Internal Identifier"
        ),
    },
    "Workers": {
        "label": "Workers",
        "endpoint_key": "workers",
        "description": (
            "Expected columns: First Name, Last Name, Job Title, Job Title ID, DOB, "
            "Street Address, Suburb, Postcode, State, Country, Internal Id, Project ID, Employer ID"
        ),
    },
}

DISPLAY_TO_SHEET = {cfg["label"]: sheet_name for sheet_name, cfg in SHEET_CONFIG.items()}


def get_token(auth_endpoint: str, email: str, password: str, tenant: str) -> str:
    headers = {"accept": "application/json", "Content-Type": "application/json"}
    body = {"email": email, "password": password, "tenant": tenant}
    response = requests.post(auth_endpoint, headers=headers, json=body, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    data = response.json()
    if "token" not in data:
        raise ValueError("No token found in response. Check credentials / tenant.")
    return data["token"]


def post_to_api(token: str, payload: dict, endpoint: str) -> Tuple[int, str]:
    headers = {
        "Authorization": "Bearer " + token,
        "accept": "application/json",
        "Content-Type": "application/json",
    }
    response = requests.post(endpoint, headers=headers, json=payload, timeout=REQUEST_TIMEOUT)
    return response.status_code, response.text


def get_all_project_ids(token: str, projects_endpoint: str) -> List[str]:
    """Fetch all project IDs by paginating through the projects endpoint (take=100 per page)."""
    headers = {
        "Authorization": "Bearer " + token,
        "accept": "application/json",
    }
    all_ids: List[str] = []
    skip = 0
    take = 100

    while True:
        response = requests.get(
            projects_endpoint,
            headers=headers,
            params={"skip": skip, "take": take},
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        data = response.json()

        # Handle both list responses and paginated object responses
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            items = data.get("items") or data.get("data") or data.get("results") or []
        else:
            break

        for item in items:
            if isinstance(item, dict):
                project_id = item.get("projectId") or item.get("id")
                if project_id:
                    all_ids.append(str(project_id))

        if len(items) < take:
            break
        skip += take

    return all_ids


def post_json_to_api(token: str, payload: dict, endpoint: str) -> Tuple[int, Any]:
    headers = {
        "Authorization": "Bearer " + token,
        "accept": "application/json",
        "Content-Type": "application/json",
    }
    response = requests.post(endpoint, headers=headers, json=payload, timeout=REQUEST_TIMEOUT)
    content_type = response.headers.get("Content-Type", "")
    if "application/json" in content_type.lower() and response.text:
        try:
            parsed = response.json()
        except ValueError:
            parsed = response.text
    else:
        parsed = response.text
    return response.status_code, parsed


def workbook_sheet_options(file_bytes: bytes) -> Tuple[openpyxl.Workbook, List[str], List[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    available_display_names = []
    missing_template_sheets = []

    for sheet_name, cfg in SHEET_CONFIG.items():
        if sheet_name in wb.sheetnames:
            available_display_names.append(cfg["label"])
        else:
            missing_template_sheets.append(sheet_name)

    return wb, available_display_names, missing_template_sheets


def normalize_date(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(cleaned, fmt).date().isoformat()
            except ValueError:
                continue
        raise ValueError(f"Unsupported DOB format: {value}")
    raise ValueError(f"Unsupported DOB value: {value!r}")


def extract_created_id(payload: Any) -> Optional[str]:
    if isinstance(payload, dict):
        candidate_paths = [
            payload.get("createdEntityId"),
            payload.get("id"),
            payload.get("workerProfile", {}).get("id") if isinstance(payload.get("workerProfile"), dict) else None,
            payload.get("data", {}).get("id") if isinstance(payload.get("data"), dict) else None,
            payload.get("result", {}).get("id") if isinstance(payload.get("result"), dict) else None,
        ]
        for candidate in candidate_paths:
            if candidate:
                return str(candidate)
    return None


def build_user_payload(
    row: tuple, all_project_ids: Optional[List[str]] = None
) -> Tuple[bool, dict]:
    email_cell = row[0] if len(row) > 0 else None
    name = row[1] if len(row) > 1 else None
    mobile = row[2] if len(row) > 2 else None
    title = row[3] if len(row) > 3 else None
    internal_identifier = row[4] if len(row) > 4 else None
    user_project_ids_raw = row[5] if len(row) > 5 else None

    if email_cell is None:
        return False, {}

    role_names = ["safetymanager"]
    is_admin = any(r.lower() == "admin" for r in role_names)

    # Resolve project IDs — treat "All" / "ALL" (case-insensitive) as assign-to-all-projects,
    # but skip that expansion for Admin users (they have implicit access).
    is_all = str(user_project_ids_raw).strip().lower() == "all" if user_project_ids_raw is not None else False
    if is_all and not is_admin:
        project_ids = all_project_ids or []
    elif user_project_ids_raw is not None and not is_all:
        project_ids = [user_project_ids_raw]
    else:
        project_ids = []

    payload = {
        "name": name or "",
        "title": title or "",
        "mobile": str(mobile) if mobile is not None else "",
        "email": email_cell or "",
        "internalIdentifier": str(internal_identifier) if internal_identifier is not None else "",
        "roleNames": role_names,
        "userProjectIds": project_ids,
    }
    return True, payload


def build_project_payload(row: tuple) -> Tuple[bool, dict]:
    project_name = row[0] if len(row) > 0 else None
    country = row[1] if len(row) > 1 else None
    site_address = row[2] if len(row) > 2 else None
    time_zone_string = row[3] if len(row) > 3 else None
    state = row[4] if len(row) > 4 else None
    internal_id = row[5] if len(row) > 5 else None
    region_id = row[6] if len(row) > 6 else None

    if not any([project_name, country, site_address, time_zone_string, state, internal_id, region_id]):
        return False, {}

    payload = {
        "isArchived": False,
        "name": project_name,
        "siteAddress": site_address,
        "regionId": region_id,
        "state": state,
        "timeZoneString": time_zone_string,
        "country": country,
        "internalIdentifier": internal_id,
        "siteTiming": [
            {"dayOfWeek": "Wednesday", "startTime": "07:00:00", "endTime": "17:00:00"},
            {"dayOfWeek": "Thursday", "startTime": "07:00:00", "endTime": "17:00:00"},
            {"dayOfWeek": "Friday", "startTime": "07:00:00", "endTime": "17:00:00"},
            {"dayOfWeek": "Saturday", "startTime": "07:00:00", "endTime": "17:00:00"},
            {"dayOfWeek": "Sunday", "startTime": "07:00:00", "endTime": "17:00:00"},
            {"dayOfWeek": "Monday", "startTime": "07:00:00", "endTime": "17:00:00"},
            {"dayOfWeek": "Tuesday", "startTime": "07:00:00", "endTime": "17:00:00"},
        ],
    }
    return True, payload


def build_employer_profile_payload(row: tuple) -> Tuple[bool, dict]:
    business_name = row[0] if len(row) > 0 else None
    abn = row[1] if len(row) > 1 else None
    street_address = row[2] if len(row) > 2 else None
    city = row[3] if len(row) > 3 else None
    state = row[4] if len(row) > 4 else None
    postal_code = row[5] if len(row) > 5 else None
    country = row[6] if len(row) > 6 else None
    internal_identifier = row[7] if len(row) > 7 else None

    if not any([
        business_name,
        abn,
        street_address,
        city,
        state,
        postal_code,
        country,
        internal_identifier,
    ]):
        return False, {}

    payload = {
        "businessName": business_name or "",
        "abn": str(abn) if abn is not None else "",
        "addresses": [
            {
                "addressType": "Physical",
                "streetAddress": street_address or "",
                "suburb": city or "",
                "state": state or "",
                "postCode": str(postal_code) if postal_code is not None else "",
                "country": country or "",
            }
        ],
        "internalIdentifier": str(internal_identifier) if internal_identifier is not None else "",
    }
    return True, payload


def build_worker_profile_payload(row: tuple) -> Tuple[bool, dict]:
    first_name = row[0] if len(row) > 0 else None
    last_name = row[1] if len(row) > 1 else None
    job_title = row[2] if len(row) > 2 else None
    job_title_id = row[3] if len(row) > 3 else None
    dob = row[4] if len(row) > 4 else None
    street_address = row[5] if len(row) > 5 else None
    suburb = row[6] if len(row) > 6 else None
    postcode = row[7] if len(row) > 7 else None
    state = row[8] if len(row) > 8 else None
    country = row[9] if len(row) > 9 else None
    internal_id = row[10] if len(row) > 10 else None
    project_id = row[11] if len(row) > 11 else None
    employer_id = row[12] if len(row) > 12 else None

    if not any([
        first_name,
        last_name,
        job_title,
        job_title_id,
        dob,
        street_address,
        suburb,
        postcode,
        state,
        country,
        internal_id,
        project_id,
        employer_id,
    ]):
        return False, {}

    profile_payload = {
        "firstName": str(first_name).strip() if first_name is not None else "",
        "lastName": str(last_name).strip() if last_name is not None else "",
        "jobTitle": str(job_title).strip() if job_title is not None else "",
        "jobTitleId": str(job_title_id).strip() if job_title_id is not None else "",
        "dateOfBirth": normalize_date(dob),
        "streetAddress": str(street_address).strip() if street_address is not None else "",
        "suburb": str(suburb).strip() if suburb is not None else "",
        "postCode": str(postcode).strip() if postcode is not None else "",
        "state": str(state).strip() if state is not None else "",
        "country": str(country).strip() if country is not None else "",
        "internalIdentifier": str(internal_id).strip() if internal_id is not None else "",
        "preferredCommunicationLanguage": DEFAULT_LANGUAGE,
    }
    profile_payload = {k: v for k, v in profile_payload.items() if v not in (None, "")}
    return True, profile_payload


def build_worker_assignment_payload(row: tuple, worker_profile_id: str) -> dict:
    employer_id = row[12] if len(row) > 12 else None
    project_id = row[11] if len(row) > 11 else None
    return {
        "employerId": str(employer_id).strip() if employer_id is not None else "",
        "projectId": str(project_id).strip() if project_id is not None else "",
        "workerProfileId": worker_profile_id,
    }


def process_standard_sheet(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    token: str,
    endpoint: str,
    start_row: int,
    progress_bar,
    log_area,
    progress_start: float,
    progress_span: float,
    logs: List[str],
    projects_endpoint: Optional[str] = None,
) -> Dict[str, int]:
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    total_rows = max(len(rows), start_row)

    processed = 0
    success = 0
    failed = 0

    # For the Users sheet, fetch all project IDs once up front so any row that
    # has "All" (case-insensitive) in the project column can be expanded at row-build time.
    all_project_ids: Optional[List[str]] = None
    if sheet_name == "Users" and projects_endpoint:
        logs.append("Users: Fetching all project IDs for potential 'All' assignments...")
        log_area.text("\n".join(logs[-25:]))
        try:
            all_project_ids = get_all_project_ids(token, projects_endpoint)
            logs.append(f"Users: Retrieved {len(all_project_ids)} project IDs.")
        except Exception as exc:
            logs.append(f"Users: Failed to fetch project IDs: {exc}")
        log_area.text("\n".join(logs[-25:]))

    payload_builder_map = {
        "Users": build_user_payload,
        "Projects": build_project_payload,
        "EmployerProfiles": build_employer_profile_payload,
    }
    payload_builder = payload_builder_map[sheet_name]

    logs.append(f"Starting sheet: {sheet_name}")
    log_area.text("\n".join(logs[-25:]))

    for i, row in enumerate(rows, start=1):
        if i < start_row:
            continue

        if sheet_name == "Users":
            should_send, payload = payload_builder(row, all_project_ids=all_project_ids)
        else:
            should_send, payload = payload_builder(row)
        if not should_send:
            current_progress = progress_start + progress_span * min(i / total_rows, 1.0)
            progress_bar.progress(min(current_progress, 1.0))
            continue

        processed += 1
        logs.append(f"{sheet_name} row {i}: Sending record...")

        try:
            status_code, response_text = post_to_api(token, payload, endpoint)
            if 200 <= status_code < 300:
                logs.append(f"{sheet_name} row {i}: Success (HTTP {status_code}).")
                success += 1
            else:
                logs.append(
                    f"{sheet_name} row {i}: Failed (HTTP {status_code}). Response: {response_text}"
                )
                failed += 1
        except Exception as exc:
            logs.append(f"{sheet_name} row {i}: Error sending record: {exc}")
            failed += 1

        current_progress = progress_start + progress_span * min(i / total_rows, 1.0)
        progress_bar.progress(min(current_progress, 1.0))
        log_area.text("\n".join(logs[-25:]))

    logs.append(f"Finished sheet: {sheet_name}")
    log_area.text("\n".join(logs[-25:]))

    return {"processed": processed, "success": success, "failed": failed}


def process_workers_sheet(
    workbook: openpyxl.Workbook,
    token: str,
    worker_profiles_endpoint: str,
    workers_endpoint: str,
    start_row: int,
    progress_bar,
    log_area,
    progress_start: float,
    progress_span: float,
    logs: List[str],
) -> Dict[str, int]:
    sheet = workbook["Workers"]
    rows = list(sheet.iter_rows(values_only=True))
    total_rows = max(len(rows), start_row)

    processed = 0
    success = 0
    failed = 0

    logs.append("Starting sheet: Workers")
    log_area.text("\n".join(logs[-25:]))

    for i, row in enumerate(rows, start=1):
        if i < start_row:
            continue

        try:
            should_send, profile_payload = build_worker_profile_payload(row)
            if not should_send:
                current_progress = progress_start + progress_span * min(i / total_rows, 1.0)
                progress_bar.progress(min(current_progress, 1.0))
                continue
        except Exception as exc:
            processed += 1
            failed += 1
            logs.append(f"Workers row {i}: Invalid data. {exc}")
            current_progress = progress_start + progress_span * min(i / total_rows, 1.0)
            progress_bar.progress(min(current_progress, 1.0))
            log_area.text("\n".join(logs[-25:]))
            continue

        processed += 1
        worker_name = " ".join([
            str(row[0]).strip() if len(row) > 0 and row[0] else "",
            str(row[1]).strip() if len(row) > 1 and row[1] else "",
        ]).strip()
        logs.append(f"Workers row {i}: Creating worker profile for {worker_name or 'worker'}...")

        try:
            profile_status, profile_response = post_json_to_api(token, profile_payload, worker_profiles_endpoint)
            if not (200 <= profile_status < 300):
                logs.append(
                    f"Workers row {i}: Worker profile failed (HTTP {profile_status}). Response: {profile_response}"
                )
                failed += 1
                current_progress = progress_start + progress_span * min(i / total_rows, 1.0)
                progress_bar.progress(min(current_progress, 1.0))
                log_area.text("\n".join(logs[-25:]))
                continue

            worker_profile_id = extract_created_id(profile_response)
            if not worker_profile_id:
                logs.append(
                    f"Workers row {i}: Worker profile created but no id found in response: {profile_response}"
                )
                failed += 1
                current_progress = progress_start + progress_span * min(i / total_rows, 1.0)
                progress_bar.progress(min(current_progress, 1.0))
                log_area.text("\n".join(logs[-25:]))
                continue

            worker_payload = build_worker_assignment_payload(row, worker_profile_id)
            logs.append("Workers row {i}: Assigning worker profile to project/employer...".format(i=i))
            worker_status, worker_response = post_json_to_api(token, worker_payload, workers_endpoint)
            if 200 <= worker_status < 300:
                logs.append(f"Workers row {i}: Success (profile + worker created).")
                success += 1
            else:
                logs.append(
                    f"Workers row {i}: Worker assignment failed (HTTP {worker_status}). Response: {worker_response}"
                )
                failed += 1
        except Exception as exc:
            logs.append(f"Workers row {i}: Error sending record: {exc}")
            failed += 1

        current_progress = progress_start + progress_span * min(i / total_rows, 1.0)
        progress_bar.progress(min(current_progress, 1.0))
        log_area.text("\n".join(logs[-25:]))

    logs.append("Finished sheet: Workers")
    log_area.text("\n".join(logs[-25:]))

    return {"processed": processed, "success": success, "failed": failed}


st.set_page_config(page_title="HammerTech Uploader", layout="wide")
st.title("HammerTech Uploader")

st.markdown(
    """
Upload the combined Excel template, enter your HammerTech credentials and tenant, then
choose which template sheets to process and upload. Currently supports Users, Projects, Employer Profiles, and Workers.
"""
)

st.header("Region")
region = st.selectbox(
    "Select your HammerTech region",
    ["North America", "Asia/Australia/NZ", "Europe/UK"],
)
st.caption("This controls which auth and API endpoints are used (US, AU, or EU environments).")

auth_endpoint, api_base = get_endpoints(region)
ENDPOINTS = {
    "users": f"{api_base}/users",
    "projects": f"{api_base}/projects",
    "employer_profiles": f"{api_base}/EmployerProfiles",
    "worker_profiles": f"{api_base}/WorkerProfiles",
    "workers": f"{api_base}/workers",
}

st.header("API Credentials")
col1, col2 = st.columns(2)

with col1:
    email = st.text_input("Email")
    tenant = st.text_input("Tenant")

with col2:
    password = st.text_input("Password", type="password")

st.header("Excel File")

with st.expander("Upload Template"):
    st.write(
        "Download the combined template, complete the sheets you want to upload, then upload that single workbook below."
    )
    try:
        with open("UploadTemplate.xlsx", "rb") as f:
            st.download_button(
                label="Download Combined Upload Template",
                data=f,
                file_name="UploadTemplate.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except FileNotFoundError:
        st.error("Upload template file not found on the server. Please contact the administrator.")

uploaded_file = st.file_uploader("Choose Excel file", type=["xlsx", "xlsm"])
start_row = st.number_input(
    "Start row (1-based, usually 2 to skip header)",
    min_value=1,
    value=2,
)

selected_display_sheets: List[str] = []
workbook_bytes = None

if uploaded_file is not None:
    try:
        workbook_bytes = uploaded_file.getvalue()
        preview_workbook, available_display_sheets, missing_template_sheets = workbook_sheet_options(workbook_bytes)
        del preview_workbook

        if available_display_sheets:
            st.success(f"Workbook loaded. Template sheets found: {', '.join(available_display_sheets)}")
            selected_display_sheets = st.multiselect(
                "Which sheets do you want to read and upload?",
                options=available_display_sheets,
                default=available_display_sheets,
            )

            for display_name in available_display_sheets:
                template_sheet_name = DISPLAY_TO_SHEET[display_name]
                st.caption(f"{display_name}: {SHEET_CONFIG[template_sheet_name]['description']}")
        else:
            st.error(
                "None of the expected template sheets were found. "
                f"Expected sheet names: {', '.join(SHEET_CONFIG.keys())}"
            )

        if missing_template_sheets:
            st.info("These template sheets were not found in the workbook: " + ", ".join(missing_template_sheets))
    except Exception as exc:
        st.error(f"Failed to inspect workbook: {exc}")

run_button = st.button("Run Upload")

if run_button:
    if not (email and password and tenant and uploaded_file):
        st.error("Please fill in all credential fields and upload a file.")
        st.stop()

    if not selected_display_sheets:
        st.error("Please select at least one sheet to upload.")
        st.stop()

    try:
        workbook_bytes = workbook_bytes or uploaded_file.getvalue()
        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    except Exception as exc:
        st.error(f"Failed to load workbook: {exc}")
        st.stop()

    selected_sheet_names = [DISPLAY_TO_SHEET[name] for name in selected_display_sheets]
    missing_selected_sheets = [sheet_name for sheet_name in selected_sheet_names if sheet_name not in workbook.sheetnames]
    if missing_selected_sheets:
        st.error("The uploaded workbook is missing these selected sheets: " + ", ".join(missing_selected_sheets))
        st.stop()

    try:
        with st.spinner(f"Authenticating with HammerTech ({region})..."):
            token = get_token(auth_endpoint, email, password, tenant)
        st.success("Authentication successful.")
    except Exception as exc:
        st.error(f"Authentication failed: {exc}")
        st.stop()

    logs: List[str] = []
    overall_processed = 0
    overall_success = 0
    overall_failed = 0
    per_sheet_results = []

    progress_bar = st.progress(0.0)
    log_area = st.empty()

    total_selected = len(selected_sheet_names)
    for idx, sheet_name in enumerate(selected_sheet_names):
        config = SHEET_CONFIG[sheet_name]
        progress_start = idx / total_selected
        progress_span = 1 / total_selected

        if sheet_name == "Workers":
            sheet_result = process_workers_sheet(
                workbook=workbook,
                token=token,
                worker_profiles_endpoint=ENDPOINTS["worker_profiles"],
                workers_endpoint=ENDPOINTS["workers"],
                start_row=start_row,
                progress_bar=progress_bar,
                log_area=log_area,
                progress_start=progress_start,
                progress_span=progress_span,
                logs=logs,
            )
        else:
            endpoint = ENDPOINTS[config["endpoint_key"]]
            sheet_result = process_standard_sheet(
                workbook=workbook,
                sheet_name=sheet_name,
                token=token,
                endpoint=endpoint,
                start_row=start_row,
                progress_bar=progress_bar,
                log_area=log_area,
                progress_start=progress_start,
                progress_span=progress_span,
                logs=logs,
                projects_endpoint=ENDPOINTS["projects"] if sheet_name == "Users" else None,
            )

        overall_processed += sheet_result["processed"]
        overall_success += sheet_result["success"]
        overall_failed += sheet_result["failed"]
        per_sheet_results.append(
            {
                "Sheet": config["label"],
                "Processed": sheet_result["processed"],
                "Successful": sheet_result["success"],
                "Failed": sheet_result["failed"],
            }
        )

    progress_bar.progress(1.0)

    st.success("Upload complete.")
    st.write(f"**Total rows processed:** {overall_processed}")
    st.write(f"Successful: {overall_success}")
    st.write(f"Failed: {overall_failed}")

    if per_sheet_results:
        st.subheader("Per-sheet summary")
        st.table(per_sheet_results)

    with st.expander("View full log"):
        st.text("\n".join(logs))
